"""Local Pulse employee XLSX search and analytics."""

from __future__ import annotations

import html
import json
import logging
import os
import pathlib
import re
import threading
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from typing import Any, Iterable

from ouroboros.paths import get_data_dir
from ouroboros.tools.registry import ToolContext, ToolEntry

log = logging.getLogger(__name__)

PULSE_XLSX_ENV = "OUROBOROS_PULSE_XLSX"
PULSE_DEFAULT_FILENAME = "emplo.xlsx"

RESTRICTED_MESSAGE = "Это конфиденциальная информация, она доступна только по специальному запросу."
OUT_OF_SCOPE_MESSAGE = "Эта информация содержится в других системах Банка. Я пока умею работать только с Пульс."
SUCCESSOR_OUT_OF_SCOPE_MESSAGE = (
    "В Пульсе нет данных о ваших подчиненных/приемниках в текущем файле. "
    "Эта информация может быть в другой системе Банка."
)

_STORE_CACHE: dict[tuple[str, float], "PulseDataStore"] = {}
_CACHE_LOCK = threading.Lock()

_PROFILE_SHEET = "profile"
_SPIDER_SBERQ_SHEET = "spider_sberq_focus"
_SPIDER_SBERTESTS_SHEET = "spider_sbertests"
_SPIDER_EXPERIENCE_SHEET = "spider_experience"
_SPIDER_CONCLUSIONS_SHEET = "spider_conclusions"
_TRAININGS_SHEET = "trainings"
_RATINGS_SHEET = "ratings"
_GOALS_SHEET = "goals"
_GRADE_CHANGES_SHEET = "grade_changes"

_SHEET_KIND_LABELS = {
    _PROFILE_SHEET: "профиль",
    _SPIDER_SBERQ_SHEET: "Паутинка (sberq и фокусные)",
    _SPIDER_SBERTESTS_SHEET: "Паутинка (сбертесты)",
    _SPIDER_EXPERIENCE_SHEET: "Паутинка (опыт)",
    _SPIDER_CONCLUSIONS_SHEET: "паутинка (выводы)",
    _TRAININGS_SHEET: "обучение за 2 года",
    _RATINGS_SHEET: "оценки за 5 лет",
    _GOALS_SHEET: "цели 2026",
    _GRADE_CHANGES_SHEET: "изменения грейдов за 5 лет",
}

_RESTRICTED_PATTERNS = (
    "дет",
    "семь",
    "семейн",
    "супруг",
    "супруга",
    "доход",
    "зарплат",
    "преми",
    "компенсац",
    "адрес",
    "телефон",
    "паспорт",
    "документ",
    "медицин",
    "здоров",
)

_OUT_OF_SCOPE_PATTERNS = (
    "комитет",
    "jira",
    "джира",
    "почт",
    "calendar",
    "календар",
    "доступ",
    "согласован",
    "комитетах",
    "комитеты",
)

_SUCCESSOR_PATTERNS = (
    "подчинен",
    "приемник",
    "преемник",
    "successor",
)

_EDUCATION_HINTS = (
    "заканчивал",
    "заканчивала",
    "окончил",
    "окончила",
    "учил",
    "учился",
    "училась",
    "образован",
    "вуз",
    "университет",
    "институт",
)

_INTEREST_HINTS = (
    "увлека",
    "хобби",
    "интерес",
)

_TRAINING_HINTS = (
    "курс",
    "обучен",
    "прошел",
    "прошёл",
    "прошла",
)

_RATING_HINTS = (
    "оценк",
    "квартал",
    "4 квартал",
    "последн",
)

_PEOPLE_CENTRICITY_HINTS = (
    "человекоцентр",
    "клиентоцентр",
    "забот",
    "поддержк",
    "команд",
    "отношен",
    "альтру",
    "люд",
)

_GOALS_HINTS = (
    "цел",
    "амбициозн",
    "материальн",
)

_AMBITION_KEYWORDS = (
    "стратег",
    "масштаб",
    "трансформац",
    "ai",
    "ии",
    "платформ",
    "внедр",
    "рост",
    "эффектив",
    "выручк",
    "клиентск",
    "автоматизац",
    "экосистем",
    "федерал",
    "международ",
    "критич",
    "ключев",
)

_PEOPLE_CENTRICITY_FIELDS_SBERQ = (
    "Забота и поддержка",
    "Формирование сильной команды",
    "Поощрение индивидуальности и мотивация",
    "Развитие эффективности команды",
    "Командность",
    "Построение отношений",
    "Видение клиента во всем многообразии ролей",
    "Создание долгосрочной ценности",
    "Проектирование будущего опыта",
)

_PEOPLE_CENTRICITY_FIELDS_SBERTESTS = (
    "Ориентация на поддержку других, альтруизм",
    "Склонность к позитивному восприятию людей",
    "Сотрудничать с командой",
    "Приносить пользу",
    "Эмоциональный интеллект",
)

_STOPWORDS = frozenset(
    {
        "а",
        "без",
        "был",
        "были",
        "быть",
        "в",
        "во",
        "все",
        "всех",
        "где",
        "его",
        "ее",
        "её",
        "еще",
        "ещё",
        "же",
        "за",
        "из",
        "или",
        "и",
        "какая",
        "какие",
        "какой",
        "кто",
        "который",
        "мне",
        "на",
        "о",
        "он",
        "она",
        "они",
        "по",
        "покажи",
        "покажи",
        "пока",
        "про",
        "с",
        "со",
        "тем",
        "там",
        "того",
        "у",
        "уже",
        "что",
        "этот",
        "этого",
        "этом",
        "кто",
    }
)

_LETTER_MAP = {
    "A": 5,
    "B": 4,
    "C": 3,
    "D": 2,
    "E": 1,
}

_CYRILLIC_TO_LATIN_GRADES = str.maketrans(
    {
        "А": "A",
        "В": "B",
        "С": "C",
        "Д": "D",
        "Е": "E",
        "а": "A",
        "в": "B",
        "с": "C",
        "д": "D",
        "е": "E",
    }
)


def _json_response(payload: dict[str, Any]) -> str:
    return json.dumps(payload, ensure_ascii=False, indent=2)


def _normalize_text(value: Any) -> str:
    text = str(value or "")
    text = html.unescape(text).replace("ё", "е").replace("Ё", "Е")
    text = re.sub(r"[^\w\s]+", " ", text, flags=re.UNICODE)
    text = re.sub(r"_+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def _normalize_header(value: Any) -> str:
    return _normalize_text(value)


def _tokenize(value: Any) -> list[str]:
    return [token for token in _normalize_text(value).split() if token]


def _safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _safe_int(value: Any) -> int | None:
    numeric = _safe_float(value)
    return int(numeric) if numeric is not None else None


def _clean_html_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    text = str(value)
    text = re.sub(r"(?i)<br\s*/?>", "\n", text)
    text = re.sub(r"<[^>]+>", "", text)
    text = html.unescape(text).replace("\xa0", " ")
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _json_safe_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        return int(value) if value.is_integer() else round(value, 4)
    if isinstance(value, str):
        return _clean_html_text(value)
    return value


def _normalize_tab_number(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value).rstrip("0").rstrip(".")
    text = _clean_html_text(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def _parse_quarter(value: Any) -> int:
    if value in (None, ""):
        return 0
    direct = _safe_int(value)
    if direct is not None:
        return direct if 1 <= direct <= 4 else 0
    text = _clean_html_text(value).strip().upper()
    match = re.search(r"([1-4])", text)
    if not match:
        return 0
    return int(match.group(1))


def _format_quarter_label(value: Any) -> str:
    quarter = _parse_quarter(value)
    return f"Q{quarter}" if quarter else _clean_html_text(value)


def _parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean_html_text(value)
    if not text:
        return None
    try:
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}( \d{2}:\d{2}:\d{2})?", text):
            return datetime.fromisoformat(text.replace(" ", "T")).date()
        if re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", text):
            return datetime.strptime(text, "%d.%m.%Y").date()
    except ValueError:
        return None
    return None


def _split_values(value: Any, pattern: str = r"[;\n,]+") -> list[str]:
    text = _clean_html_text(value)
    if not text:
        return []
    return [part.strip() for part in re.split(pattern, text) if part.strip()]


def _extract_query_terms(query: str, extra_stopwords: Iterable[str] = ()) -> list[str]:
    blocked = _STOPWORDS | {token for item in extra_stopwords for token in _tokenize(item)}
    return [token for token in _tokenize(query) if token not in blocked and len(token) > 1]


def _contains_any(text: str, patterns: Iterable[str]) -> bool:
    return any(pattern in text for pattern in patterns)


def _normalize_grade_letter(letter: Any) -> str:
    raw = _clean_html_text(letter).strip()
    if not raw:
        return ""
    raw = raw.translate(_CYRILLIC_TO_LATIN_GRADES).upper()
    return raw[:1]


def grade_to_score(letter: Any) -> int | None:
    return _LETTER_MAP.get(_normalize_grade_letter(letter))


def weighted_rating(values_letter: Any, result_letter: Any) -> float | None:
    values_score = grade_to_score(values_letter)
    result_score = grade_to_score(result_letter)
    if values_score is None or result_score is None:
        return None
    return round(0.3 * values_score + 0.7 * result_score, 4)


def _canonical_sheet_kind(title: str) -> str | None:
    normalized = _normalize_header(title)
    if normalized == "профиль":
        return _PROFILE_SHEET
    if normalized == "паутинка sberq и фокусные":
        return _SPIDER_SBERQ_SHEET
    if normalized == "паутинка сбертесты":
        return _SPIDER_SBERTESTS_SHEET
    if normalized == "паутинка опыт":
        return _SPIDER_EXPERIENCE_SHEET
    if normalized == "паутинка выводы":
        return _SPIDER_CONCLUSIONS_SHEET
    if normalized == "обучение за 2 года":
        return _TRAININGS_SHEET
    if normalized == "оценки за 5 лет":
        return _RATINGS_SHEET
    if normalized == "цели 2026":
        return _GOALS_SHEET
    if normalized == "изменения грейдов за 5 лет":
        return _GRADE_CHANGES_SHEET
    return None


@dataclass
class PulseDataStore:
    source_path: str
    schema: dict[str, list[str]] = field(default_factory=dict)
    profiles_by_tab: dict[str, dict[str, Any]] = field(default_factory=dict)
    spider_sberq_focus: dict[str, dict[str, Any]] = field(default_factory=dict)
    spider_sbertests: dict[str, dict[str, Any]] = field(default_factory=dict)
    spider_experience: dict[str, dict[str, Any]] = field(default_factory=dict)
    spider_conclusions: dict[str, dict[str, Any]] = field(default_factory=dict)
    trainings_by_tab: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    ratings_by_tab: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    goals_by_tab: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    grade_changes_by_tab: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    profile_full_text: dict[str, str] = field(default_factory=dict)
    fio_tokens: dict[str, set[str]] = field(default_factory=dict)
    interests_tokens: dict[str, set[str]] = field(default_factory=dict)
    education_tokens: dict[str, set[str]] = field(default_factory=dict)
    training_text: dict[str, str] = field(default_factory=dict)
    goal_text: dict[str, str] = field(default_factory=dict)
    spider_text: dict[str, str] = field(default_factory=dict)
    education_items: dict[str, list[str]] = field(default_factory=dict)
    interest_items: dict[str, list[str]] = field(default_factory=dict)
    all_tabs: set[str] = field(default_factory=set)

    def finalize(self) -> None:
        self.all_tabs = (
            set(self.profiles_by_tab)
            | set(self.spider_sberq_focus)
            | set(self.spider_sbertests)
            | set(self.spider_experience)
            | set(self.spider_conclusions)
            | set(self.trainings_by_tab)
            | set(self.ratings_by_tab)
            | set(self.goals_by_tab)
            | set(self.grade_changes_by_tab)
        )
        for tab in self.all_tabs:
            profile = self.profiles_by_tab.get(tab, {})
            combined_profile = " ".join(
                f"{key} {_clean_html_text(value)}" for key, value in profile.items() if value not in ("", None)
            )
            spider_text = " ".join(
                _clean_html_text(value)
                for section in (
                    self.spider_sberq_focus.get(tab, {}),
                    self.spider_sbertests.get(tab, {}),
                    self.spider_experience.get(tab, {}),
                    self.spider_conclusions.get(tab, {}),
                )
                for value in section.values()
                if value not in ("", None)
            )
            training_text = " ".join(
                _clean_html_text(item.get("название курса", "")) for item in self.trainings_by_tab.get(tab, [])
            )
            goal_text = " ".join(
                " ".join(
                    _clean_html_text(item.get(key, ""))
                    for key in (
                        "Название цели",
                        "Описание цели",
                        "Название Ключевого результата",
                        "Описание Ключевого результата",
                    )
                )
                for item in self.goals_by_tab.get(tab, [])
            )
            self.profile_full_text[tab] = _normalize_text(
                f"{combined_profile} {training_text} {goal_text} {spider_text}"
            )
            self.fio_tokens[tab] = set(_tokenize(profile.get("ФИО", "")))
            self.interest_items[tab] = _split_values(profile.get("Интересы", ""))
            self.interests_tokens[tab] = {token for item in self.interest_items[tab] for token in _tokenize(item)}
            self.education_items[tab] = _split_values(profile.get("Образование", ""), pattern=r"[;\n]+")
            self.education_tokens[tab] = {token for item in self.education_items[tab] for token in _tokenize(item)}
            self.training_text[tab] = _normalize_text(training_text)
            self.goal_text[tab] = _normalize_text(goal_text)
            self.spider_text[tab] = _normalize_text(spider_text)

    def person_brief(self, tab: str) -> dict[str, Any]:
        profile = self.profiles_by_tab.get(tab, {})
        return {
            "ref": f"tab:{tab}",
            "tab_number": tab,
            "fio": profile.get("ФИО", "") or "",
            "role": profile.get("Роль", "") or "",
            "grade": profile.get("Грейд", "") or "",
        }

    def person_card(self, tab: str) -> dict[str, Any]:
        profile = self.profiles_by_tab.get(tab, {})
        return {
            **self.person_brief(tab),
            "profile": profile,
            "spider": {
                "sberq_focus": self.spider_sberq_focus.get(tab, {}),
                "sbertests": self.spider_sbertests.get(tab, {}),
                "experience": self.spider_experience.get(tab, {}),
                "conclusions": self.spider_conclusions.get(tab, {}),
            },
            "trainings": self.trainings_by_tab.get(tab, []),
            "ratings": _sorted_ratings(self.ratings_by_tab.get(tab, [])),
            "goals": self.goals_by_tab.get(tab, []),
            "grade_changes": _sorted_grade_changes(self.grade_changes_by_tab.get(tab, [])),
        }

    def compact_person_card(self, tab: str) -> dict[str, Any]:
        profile = self.profiles_by_tab.get(tab, {})
        return {
            **self.person_brief(tab),
            "functional_block": profile.get("Функциональный блок", ""),
            "tenure_in_role": profile.get("Стаж в должности", ""),
            "tenure_in_sber": profile.get("Стаж в Сбере", ""),
            "education": profile.get("Образование", ""),
            "interests": profile.get("Интересы", ""),
            "achievements": profile.get("Ключевые достижения", ""),
            "ratings": _sorted_ratings(self.ratings_by_tab.get(tab, []))[-4:],
            "goals": self.goals_by_tab.get(tab, [])[:5],
            "spider_highlights": {
                "sberq_focus": self.spider_sberq_focus.get(tab, {}),
                "sbertests": self.spider_sbertests.get(tab, {}),
                "conclusions": self.spider_conclusions.get(tab, {}),
            },
        }


def _append_row(bucket: dict[str, list[dict[str, Any]]], tab: str, row: dict[str, Any]) -> None:
    bucket.setdefault(tab, []).append(row)


def _sorted_ratings(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        items,
        key=lambda row: (
            _safe_int(row.get("Год")) or 0,
            _parse_quarter(row.get("Квартал")),
        ),
    )


def _sorted_grade_changes(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        items,
        key=lambda row: _parse_date(row.get("Календарный день")) or date.min,
    )


def _resolve_workbook_path() -> pathlib.Path:
    override = (os.environ.get(PULSE_XLSX_ENV, "") or "").strip()
    if override:
        return pathlib.Path(override).expanduser()
    return get_data_dir() / PULSE_DEFAULT_FILENAME


def load_pulse_data(path: pathlib.Path) -> PulseDataStore:
    resolved = path.expanduser().resolve()
    stat = resolved.stat()
    cache_key = (str(resolved), stat.st_mtime)
    with _CACHE_LOCK:
        cached = _STORE_CACHE.get(cache_key)
        if cached is not None:
            return cached

    from openpyxl import load_workbook

    workbook = load_workbook(resolved, read_only=True, data_only=True)
    store = PulseDataStore(source_path=str(resolved))
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        kind = _canonical_sheet_kind(sheet_name)
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        except StopIteration:
            continue
        if not any(headers):
            continue
        store.schema[kind or sheet_name] = headers
        normalized_headers = [_normalize_header(header) for header in headers]
        if kind is None:
            continue
        if kind == _TRAININGS_SHEET:
            tab_index = _find_header_index(normalized_headers, {"тн"})
        else:
            tab_index = _find_header_index(normalized_headers, {"табельный номер"})
        if tab_index is None:
            continue
        for row_values in rows:
            row_dict = {
                headers[idx]: _json_safe_value(row_values[idx] if idx < len(row_values) else "")
                for idx in range(len(headers))
                if headers[idx]
            }
            tab_number = _normalize_tab_number(row_values[tab_index] if tab_index < len(row_values) else "")
            if not tab_number:
                continue
            if kind == _PROFILE_SHEET:
                store.profiles_by_tab[tab_number] = row_dict
            elif kind == _SPIDER_SBERQ_SHEET:
                store.spider_sberq_focus[tab_number] = row_dict
            elif kind == _SPIDER_SBERTESTS_SHEET:
                store.spider_sbertests[tab_number] = row_dict
            elif kind == _SPIDER_EXPERIENCE_SHEET:
                store.spider_experience[tab_number] = row_dict
            elif kind == _SPIDER_CONCLUSIONS_SHEET:
                store.spider_conclusions[tab_number] = row_dict
            elif kind == _TRAININGS_SHEET:
                _append_row(store.trainings_by_tab, tab_number, row_dict)
            elif kind == _RATINGS_SHEET:
                _append_row(store.ratings_by_tab, tab_number, row_dict)
            elif kind == _GOALS_SHEET:
                _append_row(store.goals_by_tab, tab_number, row_dict)
            elif kind == _GRADE_CHANGES_SHEET:
                _append_row(store.grade_changes_by_tab, tab_number, row_dict)
    workbook.close()
    store.finalize()
    with _CACHE_LOCK:
        for old_key in [key for key in _STORE_CACHE if key[0] == str(resolved) and key != cache_key]:
            _STORE_CACHE.pop(old_key, None)
        _STORE_CACHE[cache_key] = store
    return store


def _find_header_index(normalized_headers: list[str], candidates: set[str]) -> int | None:
    for idx, header in enumerate(normalized_headers):
        if header in candidates:
            return idx
    return None


def _find_person_candidates(store: PulseDataStore, person: str, limit: int = 10) -> list[dict[str, Any]]:
    query = _clean_html_text(person).strip()
    query_norm = _normalize_text(query)
    query_tokens = set(_tokenize(query))
    if not query_norm:
        return []
    is_tab_query = query_norm.replace(" ", "").isdigit()
    candidates: list[dict[str, Any]] = []
    for tab in sorted(store.all_tabs):
        profile = store.profiles_by_tab.get(tab, {})
        fio = profile.get("ФИО", "")
        fio_norm = _normalize_text(fio)
        fio_tokens = store.fio_tokens.get(tab, set())
        score = 0.0
        reasons: list[str] = []
        if is_tab_query and _normalize_tab_number(query_norm) == tab:
            score = max(score, 100.0)
            reasons.append("exact_tab")
        if fio_norm == query_norm and query_norm:
            score = max(score, 95.0)
            reasons.append("exact_fio")
        if len(query_tokens) >= 2 and query_tokens.issubset(fio_tokens):
            score = max(score, 88.0)
            reasons.append("name_tokens")
        if query_norm and query_norm in fio_norm:
            score = max(score, 78.0)
            reasons.append("fio_substring")
        token_overlap = len(query_tokens & fio_tokens)
        if token_overlap:
            score = max(score, 55.0 + 7.0 * token_overlap)
            reasons.append("token_overlap")
        seq_score = SequenceMatcher(None, query_norm, fio_norm).ratio()
        if seq_score >= 0.6:
            score = max(score, round(seq_score * 60.0, 2))
            reasons.append("fuzzy")
        if score < 35.0:
            continue
        candidates.append(
            {
                **store.person_brief(tab),
                "score": round(score, 2),
                "match_reasons": reasons,
            }
        )
    candidates.sort(key=lambda item: (-float(item["score"]), item["fio"]))
    return candidates[: max(1, limit)]


def _find_person_response(store: PulseDataStore, person: str, limit: int = 10) -> dict[str, Any]:
    candidates = _find_person_candidates(store, person, limit)
    if not candidates:
        return {"status": "not_found", "query": person}
    if len(candidates) == 1:
        candidate = dict(candidates[0])
        candidate.pop("score", None)
        candidate.pop("match_reasons", None)
        return {"status": "found", "person": candidate}
    top, second = candidates[0], candidates[1]
    if float(top["score"]) >= 70.0 and float(top["score"]) - float(second["score"]) >= 12.0:
        candidate = dict(top)
        candidate.pop("score", None)
        candidate.pop("match_reasons", None)
        return {"status": "found", "person": candidate}
    return {"status": "ambiguous", "candidates": candidates}


def _resolve_person_tab(
    store: PulseDataStore,
    *,
    ref: str = "",
    person: str = "",
    limit: int = 10,
) -> tuple[str | None, dict[str, Any] | None]:
    clean_ref = _clean_html_text(ref).strip()
    if clean_ref:
        tab = clean_ref[4:] if clean_ref.startswith("tab:") else clean_ref
        if tab in store.all_tabs:
            return tab, None
        return None, {"status": "not_found", "ref": clean_ref}
    match = _find_person_response(store, person, limit)
    if match["status"] != "found":
        return None, match
    return str(match["person"]["tab_number"]), None


def _build_ratings_summary(
    ratings: list[dict[str, Any]],
    *,
    last_n: int = 4,
) -> dict[str, Any]:
    periods: list[dict[str, Any]] = []
    for row in _sorted_ratings(ratings)[-last_n:]:
        values_letter = row.get("Оценка за ценности", "")
        result_letter = row.get("Оценка за результат", "")
        period = {
            "year": _safe_int(row.get("Год")) or 0,
            "quarter": _parse_quarter(row.get("Квартал")),
            "values_letter": _normalize_grade_letter(values_letter),
            "values_score": grade_to_score(values_letter),
            "result_letter": _normalize_grade_letter(result_letter),
            "result_score": grade_to_score(result_letter),
            "weighted_score": weighted_rating(values_letter, result_letter),
        }
        periods.append(period)
    numeric = [item["weighted_score"] for item in periods if item["weighted_score"] is not None]
    return {
        "periods": periods,
        "average_weighted_score": round(sum(numeric) / len(numeric), 4) if numeric else None,
        "formula": "0.3 * values_score + 0.7 * result_score",
    }


def _build_trainings_summary(trainings: list[dict[str, Any]]) -> dict[str, Any]:
    today = date.today()
    cutoff = today - timedelta(days=365)
    recent: list[dict[str, Any]] = []
    future: list[dict[str, Any]] = []
    for item in trainings:
        completed = _parse_date(item.get("дата завершения"))
        entry = {
            "course_name": item.get("название курса", ""),
            "completion_date": item.get("дата завершения", ""),
        }
        if completed is None:
            continue
        if completed > today:
            future.append(entry)
        elif completed >= cutoff:
            recent.append(entry)
    recent.sort(key=lambda item: str(item["completion_date"]))
    future.sort(key=lambda item: str(item["completion_date"]))
    return {
        "count_last_year": len(recent),
        "trainings_last_year": recent,
        "future_trainings": future,
    }


def _build_people_centricity_summary(store: PulseDataStore, tab: str) -> dict[str, Any]:
    sberq = {
        field: store.spider_sberq_focus.get(tab, {}).get(field, "")
        for field in _PEOPLE_CENTRICITY_FIELDS_SBERQ
        if field in store.spider_sberq_focus.get(tab, {})
    }
    sbertests = {
        field: store.spider_sbertests.get(tab, {}).get(field, "")
        for field in _PEOPLE_CENTRICITY_FIELDS_SBERTESTS
        if field in store.spider_sbertests.get(tab, {})
    }
    conclusions = {
        key: value
        for key, value in store.spider_conclusions.get(tab, {}).items()
        if key != "табельный номер" and value not in ("", None)
    }
    numeric_values = [
        float(value)
        for value in list(sberq.values()) + list(sbertests.values())
        if _safe_float(value) is not None
        for value in [value]
    ]
    return {
        "sberq_focus": sberq,
        "sbertests": sbertests,
        "conclusions": conclusions,
        "summary": {
            "signal_fields": sorted(list(sberq)) + sorted(list(sbertests)) + sorted(list(conclusions)),
            "average_numeric_score": round(sum(numeric_values) / len(numeric_values), 4) if numeric_values else None,
        },
    }


def _goal_ambition_signals(goals: list[dict[str, Any]]) -> tuple[float, list[str], list[dict[str, Any]]]:
    total_score = 0.0
    reasons: list[str] = []
    excerpts: list[dict[str, Any]] = []
    for goal in goals:
        goal_text = " ".join(
            _clean_html_text(goal.get(key, ""))
            for key in (
                "Название цели",
                "Описание цели",
                "Название Ключевого результата",
                "Описание Ключевого результата",
            )
        )
        normalized = _normalize_text(goal_text)
        weights = [
            _safe_float(goal.get(key))
            for key in ("Вес цели Q1", "Вес цели Q2", "Вес цели Q3", "Вес цели Q4", "Вес цели Y")
        ]
        weight_values = [value for value in weights if value is not None]
        max_weight = max(weight_values) if weight_values else 0.0
        keyword_hits = sum(1 for keyword in _AMBITION_KEYWORDS if keyword in normalized)
        kr_present = 1 if _clean_html_text(goal.get("Название Ключевого результата")) else 0
        length_bonus = min(len(goal_text) / 80.0, 10.0)
        score = max_weight + keyword_hits * 8.0 + kr_present * 5.0 + length_bonus
        total_score += score
        if max_weight:
            reasons.append(f"вес цели до {int(max_weight)}")
        if keyword_hits:
            reasons.append(f"масштабные маркеры: {keyword_hits}")
        if kr_present:
            reasons.append("есть ключевой результат")
        excerpts.append(
            {
                "goal_name": goal.get("Название цели", ""),
                "goal_description": goal.get("Описание цели", ""),
                "score": round(score, 2),
            }
        )
    deduped_reasons = list(dict.fromkeys(reasons))
    excerpts.sort(key=lambda item: (-float(item["score"]), item["goal_name"]))
    return round(total_score, 2), deduped_reasons, excerpts


def _query_person(store: PulseDataStore, tab: str, query: str) -> dict[str, Any]:
    profile = store.profiles_by_tab.get(tab, {})
    query_norm = _normalize_text(query)
    result: dict[str, Any]
    matched_fields: list[str]
    intent = "general"
    if _contains_any(query_norm, _EDUCATION_HINTS):
        intent = "education"
        matched_fields = ["Образование", "Ученые степени", "Научная деятельность"]
        result = {
            "profile": {
                key: profile.get(key, "")
                for key in matched_fields
                if profile.get(key, "") not in ("", None)
            }
        }
    elif _contains_any(query_norm, _INTEREST_HINTS):
        intent = "interests"
        matched_fields = ["Интересы"]
        result = {"profile": {"Интересы": profile.get("Интересы", "")}}
    elif _contains_any(query_norm, _TRAINING_HINTS):
        intent = "trainings"
        matched_fields = ["Курсы", "обучение за 2 года"]
        result = _build_trainings_summary(store.trainings_by_tab.get(tab, []))
        profile_courses = profile.get("Курсы", "")
        if profile_courses:
            result["profile_courses"] = profile_courses
    elif _is_rating_query(query, query_norm):
        intent = "ratings"
        matched_fields = ["оценки за 5 лет"]
        result = _build_ratings_summary(store.ratings_by_tab.get(tab, []), last_n=4)
    elif _contains_any(query_norm, _PEOPLE_CENTRICITY_HINTS):
        intent = "people_centricity"
        matched_fields = ["паутинка sberq", "паутинка сбертесты", "паутинка выводы"]
        result = _build_people_centricity_summary(store, tab)
    elif _contains_any(query_norm, _GOALS_HINTS):
        intent = "goals"
        matched_fields = ["цели 2026"]
        ambition_score, reasons, excerpts = _goal_ambition_signals(store.goals_by_tab.get(tab, []))
        result = {
            "goals": store.goals_by_tab.get(tab, []),
            "ambition_signals": {
                "score": ambition_score,
                "reasons": reasons,
                "goal_excerpts": excerpts[:5],
            },
        }
    else:
        matched_fields = _find_matching_field_names(store.person_card(tab), query)
        result = {
            "profile": profile,
            "spider": store.person_card(tab)["spider"],
            "trainings": store.trainings_by_tab.get(tab, []),
            "ratings": _sorted_ratings(store.ratings_by_tab.get(tab, []))[-4:],
            "goals": store.goals_by_tab.get(tab, [])[:5],
        }
    return {
        "status": "found",
        "person": store.person_brief(tab),
        "query": query,
        "intent": intent,
        "matched_fields": matched_fields,
        "result": result,
    }


def _find_matching_field_names(data: dict[str, Any], query: str) -> list[str]:
    terms = _extract_query_terms(query)
    if not terms:
        return []
    matches: list[str] = []
    for key, value in data.items():
        if isinstance(value, dict):
            nested = _find_matching_field_names(value, query)
            if nested:
                matches.extend(f"{key}.{item}" for item in nested)
        elif isinstance(value, list):
            text = _normalize_text(" ".join(_clean_html_text(item) for item in value))
            if all(term in text for term in terms):
                matches.append(key)
        else:
            text = _normalize_text(value)
            if all(term in text for term in terms):
                matches.append(key)
    return matches


def _rating_query_letters(query: str) -> list[str]:
    letters = re.findall(r"\b([ABCDEАВСДЕ])\b", query, flags=re.IGNORECASE)
    normalized = [_normalize_grade_letter(letter) for letter in letters]
    return [letter for letter in normalized if letter in _LETTER_MAP]


def _detect_rating_type(query_norm: str) -> str:
    if "за результат" in query_norm:
        return "result"
    if "за ценност" in query_norm:
        return "values"
    if "по любой" in query_norm or "любые" in query_norm or "любая из" in query_norm or "по обеим" in query_norm:
        return "either"
    return ""


def _is_rating_query(query: str, query_norm: str) -> bool:
    return _contains_any(query_norm, _RATING_HINTS) or bool(_rating_query_letters(query))


def _search_people(store: PulseDataStore, query: str, *, ref: str = "", person: str = "", limit: int = 10) -> dict[str, Any]:
    query_norm = _normalize_text(query)
    if _contains_any(query_norm, _SUCCESSOR_PATTERNS):
        if not _schema_contains_any(store, _SUCCESSOR_PATTERNS):
            return {"status": "out_of_scope", "message": SUCCESSOR_OUT_OF_SCOPE_MESSAGE}
    if "топ 3" in query_norm and _contains_any(query_norm, _EDUCATION_HINTS):
        return _education_top(store, query)
    if "там же" in query_norm and _contains_any(query_norm, _EDUCATION_HINTS):
        tab, resolution = _resolve_person_tab(store, ref=ref, person=person, limit=limit)
        if resolution is not None:
            return resolution
        return _search_people_same_education(store, tab or "", query, limit)
    if "тем же" in query_norm and _contains_any(query_norm, _INTEREST_HINTS):
        tab, resolution = _resolve_person_tab(store, ref=ref, person=person, limit=limit)
        if resolution is not None:
            return resolution
        return _search_people_same_interests(store, tab or "", query, limit)
    if _contains_any(query_norm, _GOALS_HINTS) and "амбициозн" in query_norm:
        return _search_people_ambitious_goals(store, query, limit)
    if _is_rating_query(query, query_norm):
        return _search_people_ratings(store, query, limit)
    if _contains_any(query_norm, _EDUCATION_HINTS):
        return _search_people_by_profile_field(
            store,
            query,
            profile_field="Образование",
            reason_prefix="Образование содержит",
            extra_stopwords=_EDUCATION_HINTS,
            limit=limit,
        )
    if _contains_any(query_norm, _INTEREST_HINTS):
        return _search_people_by_profile_field(
            store,
            query,
            profile_field="Интересы",
            reason_prefix="Интересы содержат",
            extra_stopwords=_INTEREST_HINTS,
            limit=limit,
        )
    return _search_people_fulltext(store, query, limit)


def _search_people_by_profile_field(
    store: PulseDataStore,
    query: str,
    *,
    profile_field: str,
    reason_prefix: str,
    extra_stopwords: Iterable[str] = (),
    limit: int = 10,
) -> dict[str, Any]:
    query_terms = _extract_query_terms(query, extra_stopwords=(profile_field, *tuple(extra_stopwords)))
    matches: list[dict[str, Any]] = []
    for tab, profile in store.profiles_by_tab.items():
        value = _clean_html_text(profile.get(profile_field, ""))
        if not value:
            continue
        normalized = _normalize_text(value)
        if query_terms and not all(term in normalized for term in query_terms):
            continue
        matches.append(
            {
                **store.person_brief(tab),
                "reasons": [f"{reason_prefix}: {value}"],
                "matched_fields": {"profile": {profile_field: value}},
            }
        )
    matches.sort(key=lambda item: item["fio"])
    return {"status": "found", "query": query, "matches": matches[:limit]}


def _search_people_ratings(store: PulseDataStore, query: str, limit: int = 10) -> dict[str, Any]:
    query_norm = _normalize_text(query)
    letters = _rating_query_letters(query)
    rating_type = _detect_rating_type(query_norm)
    year_match = re.search(r"\b(20\d{2})\b", query_norm)
    target_year = int(year_match.group(1)) if year_match else None
    if letters and not rating_type:
        return {
            "status": "needs_clarification",
            "question": "Уточните, искать A/B по оценке за результат, по оценке за ценности или по любой из них?",
        }
    matches: list[dict[str, Any]] = []
    for tab, rows in store.ratings_by_tab.items():
        reasons: list[str] = []
        matched_rows: list[dict[str, Any]] = []
        for row in _sorted_ratings(rows):
            year = _safe_int(row.get("Год")) or 0
            quarter_label = _format_quarter_label(row.get("Квартал"))
            if target_year is not None and year != target_year:
                continue
            values_letter = _normalize_grade_letter(row.get("Оценка за ценности"))
            result_letter = _normalize_grade_letter(row.get("Оценка за результат"))
            matched = False
            if rating_type == "result":
                matched = result_letter in letters
                if matched:
                    reasons.append(f"{year} {quarter_label}: результат {result_letter}")
            elif rating_type == "values":
                matched = values_letter in letters
                if matched:
                    reasons.append(f"{year} {quarter_label}: ценности {values_letter}")
            elif rating_type == "either":
                hit_fields = []
                if values_letter in letters:
                    hit_fields.append(f"ценности {values_letter}")
                if result_letter in letters:
                    hit_fields.append(f"результат {result_letter}")
                matched = bool(hit_fields)
                if matched:
                    reasons.append(f"{year} {quarter_label}: " + ", ".join(hit_fields))
            if matched:
                matched_rows.append(row)
        if reasons:
            matches.append(
                {
                    **store.person_brief(tab),
                    "reasons": list(dict.fromkeys(reasons)),
                    "matched_fields": {"ratings": matched_rows},
                }
            )
    matches.sort(key=lambda item: item["fio"])
    return {"status": "found", "query": query, "matches": matches[:limit]}


def _search_people_same_interests(store: PulseDataStore, tab: str, query: str, limit: int = 10) -> dict[str, Any]:
    base_interests = { _normalize_text(item): item for item in store.interest_items.get(tab, []) }
    matches: list[dict[str, Any]] = []
    for other_tab in sorted(store.all_tabs):
        if other_tab == tab:
            continue
        other_interests = { _normalize_text(item): item for item in store.interest_items.get(other_tab, []) }
        overlap_keys = sorted(set(base_interests) & set(other_interests))
        if not overlap_keys:
            continue
        overlap = [base_interests[key] for key in overlap_keys if key]
        matches.append(
            {
                **store.person_brief(other_tab),
                "reasons": [f"Интересы пересекаются: {', '.join(overlap)}"],
                "matched_fields": {"profile": {"Интересы": store.profiles_by_tab.get(other_tab, {}).get("Интересы", "")}},
            }
        )
    return {"status": "found", "query": query, "matches": matches[:limit]}


def _search_people_same_education(store: PulseDataStore, tab: str, query: str, limit: int = 10) -> dict[str, Any]:
    base_items = { _normalize_text(item): item for item in store.education_items.get(tab, []) }
    matches: list[dict[str, Any]] = []
    for other_tab in sorted(store.all_tabs):
        if other_tab == tab:
            continue
        other_items = { _normalize_text(item): item for item in store.education_items.get(other_tab, []) }
        overlap_keys = sorted(set(base_items) & set(other_items))
        if not overlap_keys:
            continue
        overlap = [base_items[key] for key in overlap_keys if key]
        matches.append(
            {
                **store.person_brief(other_tab),
                "reasons": [f"Образование пересекается: {', '.join(overlap)}"],
                "matched_fields": {"profile": {"Образование": store.profiles_by_tab.get(other_tab, {}).get("Образование", "")}},
            }
        )
    return {"status": "found", "query": query, "matches": matches[:limit]}


def _search_people_ambitious_goals(store: PulseDataStore, query: str, limit: int = 10) -> dict[str, Any]:
    matches: list[dict[str, Any]] = []
    for tab, goals in store.goals_by_tab.items():
        score, reasons, excerpts = _goal_ambition_signals(goals)
        if score <= 0:
            continue
        matches.append(
            {
                **store.person_brief(tab),
                "reasons": reasons or ["Есть цели с материальными признаками"],
                "matched_fields": {"goals": excerpts[:3]},
                "ambition_score": score,
            }
        )
    matches.sort(key=lambda item: (-float(item["ambition_score"]), item["fio"]))
    return {"status": "found", "query": query, "matches": matches[:limit]}


def _education_top(store: PulseDataStore, query: str) -> dict[str, Any]:
    counts: dict[str, dict[str, Any]] = {}
    for tab, items in store.education_items.items():
        for item in items:
            normalized = _normalize_text(item)
            if not normalized:
                continue
            record = counts.setdefault(normalized, {"label": item, "count": 0, "examples": []})
            record["count"] += 1
            if len(record["examples"]) < 3:
                record["examples"].append(store.person_brief(tab))
    top = sorted(counts.values(), key=lambda item: (-int(item["count"]), item["label"]))[:3]
    return {"status": "found", "query": query, "top_institutions": top}


def _search_people_fulltext(store: PulseDataStore, query: str, limit: int = 10) -> dict[str, Any]:
    query_terms = _extract_query_terms(query)
    matches: list[dict[str, Any]] = []
    for tab in sorted(store.all_tabs):
        search_text = store.profile_full_text.get(tab, "")
        if query_terms and not all(term in search_text for term in query_terms):
            continue
        matched_fields = _find_matching_field_names(store.person_card(tab), query)
        matches.append(
            {
                **store.person_brief(tab),
                "reasons": ["Полнотекстовое совпадение по данным Pulse"],
                "matched_fields": matched_fields,
            }
        )
    return {"status": "found", "query": query, "matches": matches[:limit]}


def _schema_contains_any(store: PulseDataStore, patterns: Iterable[str]) -> bool:
    normalized_patterns = tuple(pattern.lower() for pattern in patterns)
    for headers in store.schema.values():
        for header in headers:
            normalized = _normalize_header(header)
            if any(pattern in normalized for pattern in normalized_patterns):
                return True
    return False


def _compare_people(store: PulseDataStore, people: list[str], limit: int = 10) -> dict[str, Any]:
    comparisons: list[dict[str, Any]] = []
    for item in people[:limit]:
        tab, resolution = _resolve_person_tab(store, person=item, limit=limit)
        if resolution is not None:
            comparisons.append({"query": item, "resolution": resolution})
            continue
        comparisons.append({"query": item, "person": store.compact_person_card(tab or "")})
    return {"status": "found", "people": comparisons}


def _schema_payload(store: PulseDataStore) -> dict[str, Any]:
    return {
        "status": "ok",
        "source_path": store.source_path,
        "sheets": {
            _SHEET_KIND_LABELS.get(kind, kind): headers
            for kind, headers in sorted(store.schema.items(), key=lambda item: item[0])
        },
        "capabilities": [
            "find_person",
            "get_person",
            "query_person",
            "search_people",
            "compare_people",
            "schema",
        ],
    }


def _pulse_people_search(
    ctx: ToolContext,
    mode: str,
    query: str = "",
    person: str = "",
    people: list[str] | None = None,
    ref: str = "",
    filters: dict[str, Any] | None = None,
    limit: int = 10,
) -> str:
    del ctx
    del filters
    try:
        query_text = _clean_html_text(query)
        query_norm = _normalize_text(query_text)
        if _contains_any(query_norm, _RESTRICTED_PATTERNS):
            return _json_response({"status": "restricted", "message": RESTRICTED_MESSAGE})
        if _contains_any(query_norm, _OUT_OF_SCOPE_PATTERNS):
            return _json_response({"status": "out_of_scope", "message": OUT_OF_SCOPE_MESSAGE})

        workbook_path = _resolve_workbook_path()
        if not workbook_path.exists():
            return _json_response(
                {
                    "status": "error",
                    "message": f"Pulse XLSX file not found: {workbook_path}",
                    "path": str(workbook_path),
                }
            )
        store = load_pulse_data(workbook_path)
        if mode == "schema":
            return _json_response(_schema_payload(store))
        if mode == "find_person":
            return _json_response(_find_person_response(store, person or query_text, limit))
        if mode == "get_person":
            tab, resolution = _resolve_person_tab(store, ref=ref, person=person or query_text, limit=limit)
            if resolution is not None:
                return _json_response(resolution)
            return _json_response({"status": "found", "person": store.person_card(tab or "")})
        if mode == "query_person":
            tab, resolution = _resolve_person_tab(store, ref=ref, person=person or query_text, limit=limit)
            if resolution is not None:
                return _json_response(resolution)
            return _json_response(_query_person(store, tab or "", query_text or person))
        if mode == "search_people":
            return _json_response(_search_people(store, query_text, ref=ref, person=person, limit=limit))
        if mode == "compare_people":
            return _json_response(_compare_people(store, people or [], limit))
        return _json_response({"status": "error", "message": f"Unsupported mode: {mode}"})
    except Exception as exc:  # pragma: no cover - defensive envelope
        log.warning("pulse_people_search failed", exc_info=True)
        return _json_response({"status": "error", "message": f"pulse_people_search failed: {exc}"})


def get_tools() -> list[ToolEntry]:
    return [
        ToolEntry(
            "pulse_people_search",
            {
                "name": "pulse_people_search",
                "description": (
                    "Search and analyze local Pulse employee XLSX data. "
                    "Use for employee profiles, education, interests, trainings, "
                    "ratings, goals, spider scores, comparisons, and people matching criteria."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "mode": {
                            "type": "string",
                            "enum": [
                                "find_person",
                                "get_person",
                                "query_person",
                                "search_people",
                                "compare_people",
                                "schema",
                            ],
                        },
                        "query": {
                            "type": "string",
                            "description": "Natural-language query or search phrase.",
                        },
                        "person": {
                            "type": "string",
                            "description": "Person name, partial FIO, tab number, or ref-like text.",
                        },
                        "people": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "People to compare.",
                        },
                        "ref": {
                            "type": "string",
                            "description": "Stable ref returned by a previous call, e.g. tab:5925902.",
                        },
                        "filters": {
                            "type": "object",
                            "description": "Optional structured filters inferred by the model.",
                        },
                        "limit": {
                            "type": "integer",
                            "default": 10,
                        },
                    },
                    "required": ["mode"],
                },
            },
            _pulse_people_search,
            timeout_sec=120,
        )
    ]


__all__ = [
    "PulseDataStore",
    "_pulse_people_search",
    "grade_to_score",
    "weighted_rating",
    "load_pulse_data",
    "get_tools",
]
