from __future__ import annotations

import json
from datetime import date, timedelta

import pytest
from openpyxl import Workbook

from ouroboros.tools.registry import ToolContext


def _make_ctx(tmp_path):
    repo_dir = tmp_path / "repo"
    repo_dir.mkdir()
    drive_root = tmp_path / "drive"
    drive_root.mkdir()
    return ToolContext(repo_dir=repo_dir, drive_root=drive_root)


def _build_workbook(path):
    today = date.today()
    wb = Workbook()

    ws = wb.active
    ws.title = "профиль"
    ws.append(
        [
            "Табельный номер",
            "Функциональный блок",
            "ЦА\\ТБ\\ПЦП",
            "ФИО",
            "Роль",
            "Грейд",
            "Стаж в должности",
            "Стаж в Сбере",
            "Ключевые достижения",
            "Награды",
            "Место работы до Сбера",
            "Пол",
            "Знание языков",
            "Интересы",
            "Образование",
            "Курсы",
            "Ученые степени",
            "Научная деятельность",
            "Корпоративные активности",
            "режим работы (офис, гибрид, удаленный)",
            "Коэффициент отработанного времени 2026",
        ]
    )
    ws.append(
        [
            "1001",
            "Технологии",
            "ЦА-1",
            "Кесян Аршан Минсалахович",
            "Руководитель направления",
            "G12",
            "3 года",
            "5 лет",
            "Запустил новую платформу AI",
            "Лучший лидер",
            "Яндекс",
            "М",
            "Русский; английский",
            "бег; шахматы",
            "МФТИ; РАНХиГС",
            "AI leadership",
            "к.т.н.",
            "Публикации по AI",
            "Менторство",
            "гибрид",
            0.95,
        ]
    )
    ws.append(
        [
            "1002",
            "Розница",
            "ЦА-2",
            "Салайева Анна Викторовна",
            "Менеджер",
            "G10",
            "2 года",
            "4 года",
            "Оптимизировала процесс",
            "",
            "ВТБ",
            "Ж",
            "Русский; английский",
            "путешествия; рисование",
            "СПбГУ",
            "People management",
            "",
            "",
            "Волонтерство",
            "офис",
            1.0,
        ]
    )
    ws.append(
        [
            "1003",
            "Технологии",
            "ЦА-3",
            "Иванов Петр Сергеевич",
            "Архитектор",
            "G11",
            "4 года",
            "6 лет",
            "Сделал критичную интеграцию",
            "",
            "Тинькофф",
            "М",
            "Русский",
            "шахматы; кино",
            "МФТИ",
            "Data architecture",
            "",
            "",
            "",
            "удаленный",
            0.9,
        ]
    )
    ws.append(
        [
            "1004",
            "Операции",
            "ЦА-4",
            "Келлер Мария Олеговна",
            "Эксперт",
            "G9",
            "1 год",
            "2 года",
            "",
            "",
            "Альфа",
            "Ж",
            "Русский",
            "йога",
            "МГУ",
            "",
            "",
            "",
            "",
            "офис",
            1.0,
        ]
    )

    ws = wb.create_sheet("Паутинка (sberq и фокусные)")
    ws.append(
        [
            "табельный номер",
            "Забота и поддержка",
            "Формирование сильной команды",
            "Командность",
            "Построение отношений",
            "Видение клиента во всем многообразии ролей",
            "Создание долгосрочной ценности",
            "Проектирование будущего опыта",
        ]
    )
    ws.append(["1001", 8, 9, 8, 7, 9, 8, 8])
    ws.append(["1002", 7, 6, 7, 8, 6, 6, 5])
    ws.append(["1003", 5, 5, 6, 6, 5, 5, 5])

    ws = wb.create_sheet("Паутинка (сбертесты)")
    ws.append(
        [
            "табельный номер",
            "Ориентация на поддержку других, альтруизм",
            "Склонность к позитивному восприятию людей",
            "Сотрудничать с командой",
            "Приносить пользу",
            "Эмоциональный интеллект",
        ]
    )
    ws.append(["1001", 7, 8, 9, 8, 7])
    ws.append(["1002", 6, 7, 7, 7, 8])
    ws.append(["1003", 5, 5, 6, 5, 5])

    ws = wb.create_sheet("Паутинка (опыт)")
    ws.append(
        [
            "табельный номер",
            "Широта контекстов",
            "Масштаб влияния",
            "Извлеченные уроки, рефлексия",
            "Жизненный интеллект / практическая мудрость",
        ]
    )
    ws.append(["1001", 8, 9, 8, 8])
    ws.append(["1002", 6, 6, 6, 6])
    ws.append(["1003", 7, 7, 7, 7])

    ws = wb.create_sheet("паутинка (выводы)")
    ws.append(
        [
            "табельный номер",
            "Итоговый вывод по паутинке",
            "Выводы по опыту",
            "Выводы по уровню управления",
            "Выводы по проф.навыкам",
        ]
    )
    ws.append(
        [
            "1001",
            "<span>Сильный лидер<br/>Заботится о людях</span>",
            "Умеет делать выводы",
            "Управленческий потенциал высокий",
            "Профиль сильный",
        ]
    )
    ws.append(["1002", "Ориентирована на команду", "", "", ""])
    ws.append(["1003", "Технический эксперт", "", "", ""])

    ws = wb.create_sheet("обучение за 2 года")
    ws.append(["ТН", "название курса", "дата завершения"])
    ws.append(["1001", "Новый курс по лидерству", today - timedelta(days=100)])
    ws.append(["1001", "Старый курс по менеджменту", today - timedelta(days=500)])
    ws.append(["1001", "Будущий курс по стратегии", today + timedelta(days=30)])
    ws.append(["1002", "Фасилитация", today - timedelta(days=50)])

    ws = wb.create_sheet("оценки за 5 лет")
    ws.append(
        [
            "табельный номер",
            "Год",
            "Квартал",
            "Оценка за ценности",
            "Оценка за результат",
        ]
    )
    ws.append(["1001", 2025, "Q1", "B", "A"])
    ws.append(["1001", 2025, "Q2", "A", "A"])
    ws.append(["1001", 2025, "Q3", "C", "B"])
    ws.append(["1001", 2025, "Q4", "B", "B"])
    ws.append(["1002", 2025, "Q3", "A", "A"])
    ws.append(["1003", 2025, "Q2", "B", "C"])

    ws = wb.create_sheet("цели 2026")
    ws.append(
        [
            "Табельный Номер",
            "Название цели",
            "Описание цели",
            "Дата начала цели",
            "Дата выполнения цели",
            "Вес цели Q1",
            "Вес цели Q2",
            "Вес цели Q3",
            "Вес цели Q4",
            "Вес цели Y",
            "Название Ключевого результата",
            "Описание Ключевого результата",
        ]
    )
    ws.append(
        [
            "1001",
            "Стратегическая AI-платформа",
            "Масштабная трансформация клиентского опыта и автоматизация",
            date(2026, 1, 1),
            date(2026, 12, 31),
            30,
            30,
            20,
            20,
            100,
            "Внедрить платформу",
            "Запуск федерального решения",
        ]
    )
    ws.append(
        [
            "1002",
            "Улучшить отчеты",
            "Повысить эффективность отчетности",
            date(2026, 1, 1),
            date(2026, 6, 30),
            15,
            15,
            0,
            0,
            30,
            "Собрать требования",
            "Новый шаблон",
        ]
    )
    ws.append(
        [
            "1003",
            "Развить интеграции",
            "Ключевая инженерная задача для платформы",
            date(2026, 1, 1),
            date(2026, 10, 31),
            20,
            20,
            20,
            10,
            70,
            "Интеграции",
            "Повысить надежность",
        ]
    )

    ws = wb.create_sheet("изменения грейдов за 5 лет")
    ws.append(
        [
            "табельный номер",
            "Календарный день",
            "Был разряд сотрудника",
            "Стал разряд сотрудника",
        ]
    )
    ws.append(["1001", date(2024, 5, 1), "G11", "G12"])
    ws.append(["1002", date(2025, 2, 1), "G9", "G10"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _call_tool(ctx, **kwargs):
    from ouroboros.tools.pulse_people import _pulse_people_search

    return json.loads(_pulse_people_search(ctx, **kwargs))


@pytest.fixture
def pulse_book(tmp_path, monkeypatch):
    path = _build_workbook(tmp_path / "pulse.xlsx")
    monkeypatch.setenv("OUROBOROS_PULSE_XLSX", str(path))
    monkeypatch.delenv("OUROBOROS_APP_ROOT", raising=False)
    return path


def test_find_person_by_partial_fio(tmp_path, pulse_book):
    ctx = _make_ctx(tmp_path)

    result = _call_tool(ctx, mode="find_person", person="Кесян")

    assert result["status"] == "found"
    assert result["person"]["ref"] == "tab:1001"
    assert result["person"]["fio"] == "Кесян Аршан Минсалахович"


def test_get_person_aggregates_sections_and_strips_html(tmp_path, pulse_book):
    ctx = _make_ctx(tmp_path)

    result = _call_tool(ctx, mode="get_person", person="Кесян")

    assert result["status"] == "found"
    person = result["person"]
    assert person["profile"]["Образование"] == "МФТИ; РАНХиГС"
    assert person["trainings"]
    assert person["ratings"]
    assert person["goals"]
    assert person["grade_changes"]
    assert person["spider"]["sberq_focus"]["Забота и поддержка"] == 8
    assert "<span>" not in person["spider"]["conclusions"]["Итоговый вывод по паутинке"]
    assert "Сильный лидер" in person["spider"]["conclusions"]["Итоговый вывод по паутинке"]


def test_query_person_returns_education_fields(tmp_path, pulse_book):
    ctx = _make_ctx(tmp_path)

    result = _call_tool(ctx, mode="query_person", person="Кесян", query="Что заканчивал Кесян?")

    assert result["status"] == "found"
    assert result["intent"] == "education"
    assert "Образование" in result["matched_fields"]
    assert "МФТИ" in result["result"]["profile"]["Образование"]


def test_query_person_returns_interests(tmp_path, pulse_book):
    ctx = _make_ctx(tmp_path)

    result = _call_tool(ctx, mode="query_person", person="Салайева", query="Чем увлекается Салайева?")

    assert result["status"] == "found"
    assert result["intent"] == "interests"
    assert result["result"]["profile"]["Интересы"] == "путешествия; рисование"


def test_query_person_returns_last_four_quarter_average(tmp_path, pulse_book):
    ctx = _make_ctx(tmp_path)

    result = _call_tool(
        ctx,
        mode="query_person",
        person="Кажоян Шерафруза",
        ref="tab:1001",
        query="Какая средняя оценка за последние 4 квартала?",
    )

    assert result["status"] == "found"
    assert result["intent"] == "ratings"
    assert len(result["result"]["periods"]) == 4
    assert result["result"]["formula"] == "0.3 * values_score + 0.7 * result_score"
    assert result["result"]["average_weighted_score"] == pytest.approx(4.35, rel=1e-6)


def test_query_person_filters_trainings_for_last_year(tmp_path, pulse_book):
    ctx = _make_ctx(tmp_path)

    result = _call_tool(
        ctx,
        mode="query_person",
        person="Кесян",
        query="Сколько обучающих курсов и каких прошёл за последний год?",
    )

    assert result["status"] == "found"
    assert result["intent"] == "trainings"
    assert result["result"]["count_last_year"] == 1
    names = [item["course_name"] for item in result["result"]["trainings_last_year"]]
    assert names == ["Новый курс по лидерству"]
    assert result["result"]["future_trainings"][0]["course_name"] == "Будущий курс по стратегии"


def test_search_people_by_education(tmp_path, pulse_book):
    ctx = _make_ctx(tmp_path)

    result = _call_tool(ctx, mode="search_people", query="Покажи всех, кто окончил МФТИ")

    assert result["status"] == "found"
    refs = {match["ref"] for match in result["matches"]}
    assert refs == {"tab:1001", "tab:1003"}


def test_search_people_fulltext_does_not_crash_with_q_quarters(tmp_path, pulse_book):
    ctx = _make_ctx(tmp_path)

    result = _call_tool(ctx, mode="search_people", query="Руководитель направления")

    assert result["status"] == "found"
    assert result["matches"][0]["ref"] == "tab:1001"


def test_search_people_rating_query_needs_clarification_without_score_type(tmp_path, pulse_book):
    ctx = _make_ctx(tmp_path)

    result = _call_tool(ctx, mode="search_people", query="У кого в 2025 были оценки A?")

    assert result["status"] == "needs_clarification"
    assert "результат" in result["question"]
    assert "ценности" in result["question"]


def test_search_people_with_same_interests_uses_ref_context(tmp_path, pulse_book):
    ctx = _make_ctx(tmp_path)

    result = _call_tool(ctx, mode="search_people", ref="tab:1001", query="А кто еще увлекается тем же?")

    assert result["status"] == "found"
    refs = {match["ref"] for match in result["matches"]}
    assert "tab:1001" not in refs
    assert "tab:1003" in refs
    assert "шахматы" in " ".join(result["matches"][0]["reasons"]).lower()


def test_restricted_query_returns_restricted_status(tmp_path, pulse_book):
    ctx = _make_ctx(tmp_path)

    result = _call_tool(ctx, mode="query_person", person="Келлер", query="Сколько детей у Келлер?")

    assert result["status"] == "restricted"
    assert "конфиденциальная" in result["message"].lower()


def test_out_of_scope_query_returns_out_of_scope_status(tmp_path, pulse_book):
    ctx = _make_ctx(tmp_path)

    result = _call_tool(
        ctx,
        mode="query_person",
        person="Кесян",
        query="В каких комитетах состоит Кесян?",
    )

    assert result["status"] == "out_of_scope"
    assert "других системах банка" in result["message"].lower()


def test_missing_file_returns_error_json(tmp_path, monkeypatch):
    ctx = _make_ctx(tmp_path)
    missing = tmp_path / "does-not-exist.xlsx"
    monkeypatch.setenv("OUROBOROS_PULSE_XLSX", str(missing))

    result = _call_tool(ctx, mode="find_person", person="Кесян")

    assert result["status"] == "error"
    assert "xlsx" in result["message"].lower()


def test_default_data_dir_is_used_when_override_missing(tmp_path, monkeypatch):
    ctx = _make_ctx(tmp_path)
    monkeypatch.delenv("OUROBOROS_PULSE_XLSX", raising=False)
    monkeypatch.setenv("OUROBOROS_APP_ROOT", str(tmp_path))
    _build_workbook(tmp_path / "Data" / "emplo.xlsx")

    result = _call_tool(ctx, mode="find_person", person="Кесян")

    assert result["status"] == "found"
    assert result["person"]["ref"] == "tab:1001"
